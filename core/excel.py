import logging
import os
from typing import List

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .env import CSV_PATH, EXCEL_PATH
from .logger import setup_logging

setup_logging()
LOGGER = logging.getLogger(__name__)
HEADER_FILL = PatternFill(fill_type="solid", start_color="B8CCE4", end_color="B8CCE4")
TOTAL_FILL = PatternFill(fill_type="solid", start_color="DCE6F1", end_color="DCE6F1")
HEADER_FONT = Font(bold=True)


def apply_header_style(worksheet, start_column: int, end_column: int, row: int = 1):
    """Aplica el estilo estándar a un rango horizontal de encabezados."""
    for cells in worksheet.iter_rows(
        min_row=row,
        max_row=row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in cells:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT


def adjust_column_widths(
    worksheet, start_column: int = 1, end_column: int = None, padding: int = 2
):
    """Ajusta el ancho de cada columna al contenido de sus celdas."""
    if end_column is None:
        end_column = worksheet.max_column

    for column in worksheet.iter_cols(
        min_col=start_column, max_col=end_column, min_row=1
    ):
        max_length = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = (
            max_length + padding
        )


def convert_to_excel(file_name: str):
    if os.path.getsize(f"{CSV_PATH}/{file_name}.csv") == 0:
        LOGGER.warning(f"{file_name}.csv esta vacio no se pudo convertir a xlsx")
        return False

    # Lee el archivo especificado
    df = pd.read_csv(f"{CSV_PATH}/{file_name}.csv", sep="\t")
    df["Fecha y Hora"] = pd.to_datetime(df["Fecha y Hora"], format="%Y-%m-%d %H:%M:%S")

    with pd.ExcelWriter(f"{EXCEL_PATH}/{file_name}.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reporte")
        worksheet = writer.sheets["Reporte"]

        # Aplicar formato de fecha a la columna A
        for cell in worksheet["A"]:
            if cell.row == 1:
                continue  # encabezado
            cell.number_format = "DD/MM/YYYY HH:MM"

        apply_header_style(worksheet, 1, worksheet.max_column)
        adjust_column_widths(worksheet)

    LOGGER.info(f"{file_name}.xlsx generado exitosamente")


def merge_csv_to_excel(file_names: List[str], output_name: str):
    """Une CSV tabulados compatibles y los guarda en un único archivo Excel."""
    dataframes = []
    expected_columns = None

    for file_name in file_names:
        csv_file = f"{CSV_PATH}/{file_name}.csv"

        if not os.path.isfile(csv_file) or os.path.getsize(csv_file) == 0:
            LOGGER.warning(f"{file_name}.csv no existe o está vacío; se omitirá.")
            continue

        try:
            dataframe = pd.read_csv(csv_file, sep="\t")
        except (pd.errors.EmptyDataError, UnicodeDecodeError) as error:
            LOGGER.warning(f"No se pudo leer {file_name}.csv; se omitirá: {error}")
            continue

        if dataframe.empty:
            LOGGER.warning(f"{file_name}.csv no contiene registros; se omitirá.")
            continue

        if expected_columns is None:
            expected_columns = dataframe.columns
        elif not dataframe.columns.equals(expected_columns):
            LOGGER.error(
                f"{file_name}.csv tiene columnas distintas al primer CSV válido. "
                "No se generará el reporte consolidado."
            )
            return False

        dataframes.append(dataframe)

    if not dataframes:
        LOGGER.warning("No hay CSV con registros para consolidar.")
        return False

    dataframe = pd.concat(dataframes, ignore_index=True)
    if "Fecha y Hora" in dataframe.columns:
        dataframe["Fecha y Hora"] = pd.to_datetime(
            dataframe["Fecha y Hora"], format="%Y-%m-%d %H:%M:%S"
        )

    output_file = f"{EXCEL_PATH}/{output_name}.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Reporte")
        worksheet = writer.sheets["Reporte"]

        if "Fecha y Hora" in dataframe.columns:
            date_column = dataframe.columns.get_loc("Fecha y Hora") + 1
            for cell in worksheet.iter_cols(
                min_col=date_column, max_col=date_column, min_row=2
            ):
                for value in cell:
                    value.number_format = "DD/MM/YYYY HH:MM"

        apply_header_style(worksheet, 1, worksheet.max_column)
        adjust_column_widths(worksheet)

    LOGGER.info(f"{output_name}.xlsx consolidado exitosamente")
    return True


def tab_dinamica(file_name: str):
    # Leer datos
    df = pd.read_excel(file_name)

    # Convertir Fecha y Hora a datetime
    df["Fecha y Hora"] = pd.to_datetime(df["Fecha y Hora"])

    # Obtener únicamente la fecha (equivalente al agrupamiento por Días)
    df["Días"] = df["Fecha y Hora"].dt.date

    # Crear tabla dinámica
    pivotTable = df.groupby("Días", as_index=False).agg(
        {"Minutos": "sum", "Consumo": "sum"}
    )

    # Renombrar columnas
    pivotTable.columns = ["Etiquetas de fila", "Suma de Minutos", "Suma de Consumo"]

    # Agregar fila Total general
    grandTotal = pd.DataFrame(
        {
            "Etiquetas de fila": ["Total general"],
            "Suma de Minutos": [pivotTable["Suma de Minutos"].sum()],
            "Suma de Consumo": [pivotTable["Suma de Consumo"].sum()],
        }
    )

    finalTable = pd.concat([pivotTable, grandTotal], ignore_index=True)

    # Escribir en el mismo archivo
    with pd.ExcelWriter(
        file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay"
    ) as writer:
        # Primera hoja
        hoja = writer.book.sheetnames[0]

        # Obtener hoja para aplicar formato
        ws = writer.book[hoja]

        # Escribir desde J1
        finalTable.to_excel(
            writer,
            sheet_name=hoja,
            startrow=0,
            startcol=8,  # Columna J
            index=False,
        )

        # Encabezados (J1:L1)
        apply_header_style(ws, 9, 11)

        # Fila Total general
        ultima_fila = len(finalTable) + 1

        for celda in ws[f"I{ultima_fila}:K{ultima_fila}"][0]:
            celda.fill = TOTAL_FILL
            celda.font = HEADER_FONT

        # Ajustar ancho de columnas automáticamente
        adjust_column_widths(ws, start_column=9, end_column=11, padding=3)

    print("Tabla dinámica generada y formateada correctamente.")
